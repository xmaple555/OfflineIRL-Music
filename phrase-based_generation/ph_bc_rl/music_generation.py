import argparse
import pickle
from tqdm import tqdm
import torch

import sys
import os

project_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), '..','..'))
src_dir = os.path.join(project_dir, 'src')
sys.path.append(src_dir)
from model import Generator
from openpyxl import Workbook
from vocab import Vocab
from reward_experiment import *
from inference import inference
from utils import set_seed, network_paras

myvocab = Vocab()


def get_input_and_phrase_labels(remi_seq):
    first_section_end_idx = remi_seq.index(myvocab.token2id["Section_End"])

    first_section_seq = remi_seq[:first_section_end_idx + 1]

    phrase_labels = myvocab.extract_phrase_label(first_section_seq)

    input_bar_num = int(phrase_labels[0][1:])
    second_phrase_index = 0
    bar_num = 0
    for i in range(len(first_section_seq)):
        if "Bar" in myvocab.id2token[first_section_seq[i]]:
            bar_num += 1
        if bar_num == input_bar_num + 1:
            second_phrase_index = i
            break

    input_seq = first_section_seq[:second_phrase_index]
    return first_section_seq, input_seq, phrase_labels

if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument('--sample_num', type=int, default=10)
    parser.add_argument('--ckpt', type=str, default='')
    parser.add_argument('--max_len',
                        type=int,
                        default=2048,
                        help='number of tokens to predict')
    parser.add_argument('--t', type=float, default=1.0)
    parser.add_argument('--p', type=float, default=1.0)
    parser.add_argument('--input_song_idx', type=int, default=0)
    parser.add_argument('--input_phrase_labels', type=str, default= "A4 A4 B4 C4")
    parser.add_argument('--tempo', type=int, default=120)
    parser.add_argument('--mid_name', type=str)
    
    args = parser.parse_args()
    set_seed(0)


    with open("../ph_bc/data_pkl/test_full_seqs.pkl", "rb") as f:
        test_full_seqs = pickle.load(f)


    test_pop909_full_seqs = []
    for data in test_full_seqs:
        if data["name"] == "pop909_1":
            test_pop909_full_seqs.append(data["seq"])


    user_phrase_labels = args.input_phrase_labels.split()

    first_section_seq, input_seq, phrase_labels = get_input_and_phrase_labels(
        test_pop909_full_seqs[args.input_song_idx])
            

    assert user_phrase_labels[0] == phrase_labels[0], f"the first phrase label must be the same as the length of input sequences. first phrase label: {phrase_labels[0]}"

    music_generator = Generator(myvocab.n_tokens).cuda()
    checkpoint = torch.load(args.ckpt)

    try:
        music_generator.load_state_dict(checkpoint['generator_state_dict'])
    except:
        music_generator = Generator(829).cuda()
        music_generator.load_state_dict(checkpoint['generator_state_dict'])

    temp_file_name = '/tmp/' + str(uuid.uuid4()) + '.txt'
    experiment = RewardExperimentPhBc(temp_file_name)

    wb = Workbook()
    ws = wb.active
    ws['A1'] = 'file'
    ws['B1'] = 'melody_reward'
    ws['C1'] = 'chord_progression_reward'


    directory = f"output_music/{args.mid_name}"
    if not os.path.exists(directory):
        os.makedirs(directory)


    myvocab.REMIID2midi(input_seq,
                                os.path.join(directory, "prefix_sequence.mid"),
                                tempo=args.tempo)

    ###
    # song_idx=13
    # tmp_idx = 0
    # bar_num = 0
    # first_section_seq, input_seq, phrase_labels = get_input_and_phrase_labels(
    #     test_pop909_full_seqs[song_idx])
    
    # for i in range(len(first_section_seq)):
    #     if "Bar" in myvocab.id2token[first_section_seq[i]]:
    #         bar_num += 1
    #     if bar_num == 32 + 1:
    #         tmp_idx = i
    #         break

    # first_section_seq = first_section_seq[:tmp_idx]
    # first_section_seq += [myvocab.token2id["Section_End"]]
    # myvocab.REMIID2midi(first_section_seq,"test.mid")
    # experiment.reward_experiment_fun(first_section_seq)
    # print(experiment.rewards["melody"][-1], experiment.rewards["chord_progression"][-1])
    ###

    for i in tqdm(range(args.sample_num)):
        while True:
            sampled_remi_seq = inference(
                model=music_generator,
                myvocab=myvocab,
                max_len=args.max_len,
                input_seq=input_seq,
                input_phrase_labels=user_phrase_labels,
                t=args.t,
                p=args.p)
            if sampled_remi_seq is not None:
                break

        experiment.reward_experiment_fun(sampled_remi_seq)
        mid_name = f"{args.mid_name}_{1+i}.mid"
        ws.append([
                    mid_name,
                    experiment.rewards["melody"][-1],
                    experiment.rewards["chord_progression"][-1],
                ])

        wb.save(f"output_music/{args.mid_name}/rewards.xlsx")
        myvocab.REMIID2midi(sampled_remi_seq,
                                    os.path.join(directory, mid_name),
                                    tempo=args.tempo)


    os.remove(temp_file_name)
    ws.append([
            "average reward",
            np.mean(experiment.rewards["melody"]),
            np.mean(experiment.rewards["chord_progression"])
        ])

    wb.save(f"output_music/{args.mid_name}/rewards.xlsx")