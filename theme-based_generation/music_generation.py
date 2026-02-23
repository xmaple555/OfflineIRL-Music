import argparse
import pickle
from tqdm import tqdm
import torch

import sys
import os

project_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
src_dir = os.path.join(project_dir, 'src')
sys.path.append(src_dir)
from mymodel import myLM
from openpyxl import Workbook
from theme_preprocess.vocab import Vocab as ThemeVocab

from reward_experiment import *
from inference import inference
from utils import set_seed, network_paras

myvocab = Vocab()
mythemevocab = ThemeVocab()

def fix_ph_bc_label(words, input_phrase_labels):
    phrase_label_idx = 0
    bar_countdown = 1
    i = 0
    phrase_label = None
    while i < len(words):
        if "Bar" in myvocab.id2token[words[i]]:
            if bar_countdown == 1:
                bar_countdown = int(input_phrase_labels[phrase_label_idx][1:])
                words[i] = myvocab.token2id["Bar_" + str(bar_countdown)]
                phrase_label = myvocab.token2id["Phrase_" + input_phrase_labels[phrase_label_idx][0]]
                words.insert(i + 1, phrase_label)
                phrase_label_idx += 1
            else:
                bar_countdown = bar_countdown - 1
                words[i] = myvocab.token2id["Bar_" + str(bar_countdown)]
                words.insert(i + 1, phrase_label)
        i += 1

    assert bar_countdown == 1 and phrase_label_idx == len( input_phrase_labels) 
    return words


def convert_remi_seq(remi_seq):
    remi_seq2 = []
    for x in remi_seq:
        if mythemevocab.id2token[x] in myvocab.token2id:
            remi_seq2.append(myvocab.token2id[mythemevocab.id2token[x]])
    return remi_seq2



def get_input_and_phrase_labels(remi_seq):
    first_section_end_idx = remi_seq.index(myvocab.token2id["Section_End"])

    first_section_seq = remi_seq[:first_section_end_idx + 1]

    phrase_labels = myvocab.extract_phrase_label(first_section_seq)

    input_bar_num = int(phrase_labels[0][1:])
    second_phrase_index = 0
    bar_num = 0
    for i in range(len(first_section_seq)):
        if "Bar" in myvocab.id2token[first_section_seq[i]]:
            bar_num += 1
        if bar_num == input_bar_num + 1:
            second_phrase_index = i
            break

    input_seq = first_section_seq[:second_phrase_index]
    return first_section_seq, input_seq, phrase_labels

if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument('--sample_num', type=int, default=10)
    parser.add_argument('--t', type=float, default=1.0)
    parser.add_argument('--p', type=float, default=1.0)
    parser.add_argument('--input_song_idx', type=int, default=0)
    parser.add_argument('--bar_num', type=int, default=16)
    parser.add_argument('--tempo', type=int, default=120)
    parser.add_argument('--mid_name', type=str)
    
    args = parser.parse_args()
    set_seed(0)



    with open(os.path.join(project_dir, "phrase-based_generation/ph_bc/data_pkl",
                         "test_full_seqs.pkl"), "rb") as f:
        test_full_seqs = pickle.load(f)

    test_pop909_full_seqs = []
    for data in test_full_seqs:
        if data["name"] == "pop909_1":
            test_pop909_full_seqs.append(data["seq"])


    first_section_seq, input_seq, phrase_labels = get_input_and_phrase_labels(
        test_pop909_full_seqs[args.input_song_idx])
            

    input_seq =  myvocab.remove_ph_bc(input_seq)

    myvocab.REMIID2midi(input_seq, "/tmp/tmp_theme.mid")
    decoder_input_seq = mythemevocab.midi2REMI(
        "/tmp/tmp_theme.mid",
        trim_intro=False,
        trim_outro=False,
        theme_annotations=False,
    )
    decoder_input_seq = [
        x for x in decoder_input_seq
        if not mythemevocab.id2token[x].startswith("Tempo")
    ]



    encoder_input_seq = []
    bar_count = 0
    second_bar_idx = 0

    for i in range(len(decoder_input_seq)):
        x = decoder_input_seq[i]
        if "Bar" == mythemevocab.id2token[x]:
            bar_count += 1
            if bar_count > 2:
                second_bar_idx = i
                break
        encoder_input_seq.append(x)

    encoder_input_seq = [
        mythemevocab.token2id["Theme_Start"]
    ] + encoder_input_seq + [mythemevocab.token2id["Theme_End"]]
    decoder_input_seq.insert(second_bar_idx,
                                mythemevocab.token2id["Theme_End"])
    decoder_input_seq.insert(0, mythemevocab.token2id["Theme_Start"])

    input_tmp = args.tempo
    tmp = mythemevocab._tempo_bins[np.argmin(    abs(input_tmp - mythemevocab._tempo_bins))]
    given_tempo = mythemevocab.token2id["Tempo_{}".format(tmp)]
    decoder_input_seq = [given_tempo] + decoder_input_seq + [
        mythemevocab.token2id["Theme_Start"]
    ]


    input_phrase_labels = [f"A{args.bar_num}"]


    music_generator = myLM(mythemevocab.n_tokens,d_model=256,num_encoder_layers=6,xorpattern=[0,0,0,1,1,1]).cuda()
    music_generator.load_state_dict(torch.load("ckpt/model_ep2311.pt"))

    temp_file_name = '/tmp/' + str(uuid.uuid4()) + '.txt'
    experiment = RewardExperimentNoPhBc(temp_file_name)

    wb = Workbook()
    ws = wb.active
    ws['A1'] = 'file'
    ws['B1'] = 'melody_reward'
    ws['C1'] = 'chord_progression_reward'


    directory = f"output_music/{args.mid_name}"
    if not os.path.exists(directory):
        os.makedirs(directory)


    myvocab.REMIID2midi(input_seq,
                                os.path.join(directory, "prefix_sequence.mid"),
                                tempo=args.tempo)

    for i in tqdm(range(args.sample_num)):
        while True:
            sampled_remi_seq = inference(
                model=music_generator,
                theme_seq = encoder_input_seq,
                prompt = decoder_input_seq,
                n_bars=args.bar_num,
                t=args.t,
                p=args.p)
            

            if sampled_remi_seq is not None:
                break

        sampled_remi_seq = convert_remi_seq(sampled_remi_seq)
        sampled_remi_seq = fix_ph_bc_label(sampled_remi_seq, input_phrase_labels)
        
        experiment.reward_experiment_fun(sampled_remi_seq)

        mid_name = f"{args.mid_name}_{1+i}.mid"
        ws.append([
                    mid_name,
                    experiment.rewards["melody"][-1],
                    experiment.rewards["chord_progression"][-1],
                ])

        wb.save(f"output_music/{args.mid_name}/rewards.xlsx")
        myvocab.REMIID2midi(sampled_remi_seq,
                                    os.path.join(directory, mid_name),
                                    tempo=args.tempo)


    os.remove(temp_file_name)
    ws.append([
            "average reward",
            np.mean(experiment.rewards["melody"]),
            np.mean(experiment.rewards["chord_progression"])
        ])

    wb.save(f"output_music/{args.mid_name}/rewards.xlsx")